"""Excel导出模块 - 使用pandas和openpyxl导出Excel."""

from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict, Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelExporter:
    """Excel导出器 - 将招标信息导出为Excel文件."""
    
    def __init__(self, output_file: str = "medical_tenders.xlsx"):
        """初始化导出器.
        
        Args:
            output_file: 输出文件名
        """
        self.output_file = output_file
        
    def export(self, results: Dict[str, List[Any]], output_file: Optional[str] = None) -> str:
        """导出招标信息到Excel.
        
        Args:
            results: 按关键词分类的招标信息字典
            output_file: 输出文件名（可选）
            
        Returns:
            str: 导出的文件路径
        """
        if output_file:
            self.output_file = output_file
            
        print(f"\n正在导出到 Excel: {self.output_file}")
        
        # 准备数据
        all_data = []
        for keyword, tenders in results.items():
            for tender in tenders:
                # 清理标的物中的无效内容
                subject = tender.subject
                if subject and '公告页面' in subject:
                    subject = ''
                
                all_data.append({
                    "关键词": keyword,
                    "标题": tender.title,
                    "发布日期": tender.publish_date,
                    "公告类型": tender.notice_type,
                    "省份": tender.province,
                    "采购单位": tender.purchaser,
                    "代理机构": tender.agency,
                    "预算金额": tender.budget,
                    "标的物": subject,
                    "联系人": tender.contact_name,
                    "联系电话": tender.contact_phone,
                    "联系地址": tender.contact_address,
                    "URL": tender.url,
                })
        
        # 创建DataFrame
        df = pd.DataFrame(all_data)
        
        # 保存到Excel
        df.to_excel(self.output_file, index=False, engine='openpyxl')
        
        print(f"导出完成！共 {len(all_data)} 条记录")
        print(f"文件保存至: {Path(self.output_file).absolute()}")
        
        return self.output_file
    
    def export_multi_sheet(self, results: Dict[str, List[Any]], output_file: Optional[str] = None) -> str:
        """导出到多工作表Excel（每个关键词一个工作表）.
        
        Args:
            results: 按关键词分类的招标信息字典
            output_file: 输出文件名（可选）
            
        Returns:
            str: 导出的文件路径
        """
        if output_file:
            self.output_file = output_file
            
        print(f"\n正在导出到多工作表 Excel: {self.output_file}")
        
        # 创建工作簿
        wb = Workbook()
        wb.remove(wb.active)  # 删除默认工作表
        
        # 创建汇总工作表
        summary_ws = wb.create_sheet("汇总")
        self._create_summary_sheet(summary_ws, results)
        
        # 为每个关键词创建工作表
        for keyword, tenders in results.items():
            ws = wb.create_sheet(keyword[:31])  # Excel工作表名最多31字符
            self._create_keyword_sheet(ws, keyword, tenders)
        
        # 保存文件
        wb.save(self.output_file)
        
        total = sum(len(tenders) for tenders in results.values())
        print(f"导出完成！共 {total} 条记录，{len(results)} 个关键词")
        print(f"文件保存至: {Path(self.output_file).absolute()}")
        
        return self.output_file
    
    def _create_summary_sheet(self, ws, results: Dict[str, List[Any]]):
        """创建汇总工作表."""
        # 标题
        ws['A1'] = "医疗器械招投标信息汇总"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # 导出时间
        ws['A2'] = f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True)
        ws.merge_cells('A2:D2')
        
        # 表头
        headers = ["关键词", "记录数", "占比"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # 数据
        total = sum(len(tenders) for tenders in results.values())
        row = 5
        for keyword, tenders in results.items():
            ws.cell(row=row, column=1, value=keyword)
            ws.cell(row=row, column=2, value=len(tenders))
            ws.cell(row=row, column=3, value=f"{len(tenders)/total*100:.1f}%" if total > 0 else "0%")
            row += 1
        
        # 合计
        ws.cell(row=row, column=1, value="合计").font = Font(bold=True)
        ws.cell(row=row, column=2, value=total).font = Font(bold=True)
        ws.cell(row=row, column=3, value="100%").font = Font(bold=True)
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
    
    def _create_keyword_sheet(self, ws, keyword: str, tenders: List[Any]):
        """创建单个关键词的工作表."""
        # 标题
        ws['A1'] = f"{keyword} - 招标信息"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')
        
        # 表头
        headers = ["序号", "标题", "发布日期", "公告类型", "省份", "采购单位", "代理机构", "预算金额", "标的物", "联系人", "联系电话", "联系地址", "URL"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # 数据
        for idx, tender in enumerate(tenders, 1):
            row = idx + 3
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=tender.title)
            ws.cell(row=row, column=3, value=tender.publish_date)
            ws.cell(row=row, column=4, value=tender.notice_type)
            ws.cell(row=row, column=5, value=tender.province)
            ws.cell(row=row, column=6, value=tender.purchaser)
            ws.cell(row=row, column=7, value=tender.agency)
            ws.cell(row=row, column=8, value=tender.budget)
            ws.cell(row=row, column=9, value=tender.subject)
            ws.cell(row=row, column=10, value=tender.contact_name)
            ws.cell(row=row, column=11, value=tender.contact_phone)
            ws.cell(row=row, column=12, value=tender.contact_address)
            ws.cell(row=row, column=13, value=tender.url)
        
        # 调整列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 40  # 标的物
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 18
        ws.column_dimensions['L'].width = 35
        ws.column_dimensions['M'].width = 60
