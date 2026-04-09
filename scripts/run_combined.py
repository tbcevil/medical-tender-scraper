#!/usr/bin/env python3
"""合并CCGP和GGZY招标信息采集脚本.

将两个平台的结果导出到同一个Excel文件的不同Sheet中。

使用方法:
    python scripts/run_combined.py -k 眼科
    python scripts/run_combined.py -k 眼科 激光 -d 7 --max-results 20
"""

import sys
from pathlib import Path

# 添加src到路径
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

import argparse
import time
from datetime import datetime

from config import TenderConfig


def setup_argument_parser() -> argparse.ArgumentParser:
    """设置命令行参数解析器."""
    parser = argparse.ArgumentParser(
        description="合并CCGP和GGZY招标信息采集工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
    # 基本使用（默认搜索"眼科"）
    python scripts/run_combined.py

    # 指定关键词
    python scripts/run_combined.py -k 眼科

    # 指定多个关键词
    python scripts/run_combined.py -k 眼科 激光 显微镜

    # 指定最大结果数
    python scripts/run_combined.py --max-results 20

    # 获取所有结果
    python scripts/run_combined.py --all

    # 只抓取CCGP
    python scripts/run_combined.py --ccgp-only

    # 只抓取GGZY
    python scripts/run_combined.py --ggzy-only
        """
    )
    
    parser.add_argument(
        "-k", "--keywords",
        nargs="+",
        default=["眼科"],
        help="搜索关键词列表 (默认: 眼科)"
    )
    
    parser.add_argument(
        "-o", "--output",
        default=None,
        help="输出Excel文件名 (默认: 自动生成带日期的文件名)"
    )
    
    parser.add_argument(
        "-d", "--days",
        type=int,
        default=7,
        help="搜索最近几天的数据 (默认: 7)"
    )
    
    parser.add_argument(
        "--max-results",
        type=int,
        default=20,
        help="每个关键词最大结果数，0表示获取所有 (默认: 20)"
    )
    
    parser.add_argument(
        "--all",
        action="store_true",
        help="获取所有结果（相当于 --max-results 0）"
    )
    
    parser.add_argument(
        "--ccgp-only",
        action="store_true",
        help="只抓取CCGP数据"
    )
    
    parser.add_argument(
        "--ggzy-only",
        action="store_true",
        help="只抓取GGZY数据"
    )
    
    parser.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="显示详细输出"
    )
    
    return parser


def export_combined_excel(ccgp_results: dict, ggzy_results: dict, output_file: str) -> str:
    """导出合并的Excel文件，包含两个Sheet.
    
    Args:
        ccgp_results: CCGP抓取结果
        ggzy_results: GGZY抓取结果
        output_file: 输出文件名
        
    Returns:
        str: 导出的文件路径
    """
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    print(f"\n正在导出到 Excel: {output_file}")
    
    # 创建工作簿
    wb = Workbook()
    
    # 定义样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def create_sheet(wb, sheet_name, results):
        """创建Sheet并填充数据."""
        if sheet_name == "CCGP":
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)
        
        # 准备数据
        all_data = []
        for keyword, tenders in results.items():
            for tender in tenders:
                all_data.append({
                    "关键词": keyword,
                    "标题": tender.title,
                    "发布日期": tender.publish_date,
                    "公告类型": tender.notice_type,
                    "省份": tender.province,
                    "采购单位": tender.purchaser,
                    "代理机构": tender.agency,
                    "预算金额": tender.budget,
                    "标的物": tender.subject,
                    "联系人": tender.contact_name,
                    "联系电话": tender.contact_phone,
                    "联系地址": tender.contact_address,
                    "URL": tender.url
                })
        
        if not all_data:
            ws.append(["暂无数据"])
            return len(all_data)
        
        # 创建DataFrame
        df = pd.DataFrame(all_data)
        
        # 写入表头
        headers = list(df.columns)
        ws.append(headers)
        
        # 设置表头样式
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # 写入数据
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)
        
        # 设置列宽
        column_widths = {
            'A': 12,  # 关键词
            'B': 50,  # 标题
            'C': 12,  # 发布日期
            'D': 15,  # 公告类型
            'E': 15,  # 省份
            'F': 25,  # 采购单位
            'G': 25,  # 代理机构
            'H': 15,  # 预算金额
            'I': 40,  # 标的物
            'J': 12,  # 联系人
            'K': 15,  # 联系电话
            'L': 30,  # 联系地址
            'M': 60   # URL
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 设置数据单元格样式
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        return len(all_data)
    
    # 创建CCGP Sheet
    ccgp_count = 0
    if ccgp_results:
        ccgp_count = create_sheet(wb, "CCGP", ccgp_results)
        print(f"  CCGP Sheet: {ccgp_count} 条记录")
    
    # 创建GGZY Sheet
    ggzy_count = 0
    if ggzy_results:
        if not ccgp_results:
            wb.active.title = "GGZY"
            ggzy_count = create_sheet(wb, "GGZY", ggzy_results)
        else:
            ggzy_count = create_sheet(wb, "GGZY", ggzy_results)
        print(f"  GGZY Sheet: {ggzy_count} 条记录")
    
    # 保存文件
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    
    print(f"\n导出完成: {output_path}")
    print(f"  总记录数: {ccgp_count + ggzy_count} 条")
    
    return str(output_path)


def main():
    """主函数."""
    parser = setup_argument_parser()
    args = parser.parse_args()
    
    # 自动生成带日期的文件名
    if args.output is None:
        today = datetime.now().strftime("%Y%m%d")
        args.output = f"combined_tenders_{today}.xlsx"
    
    # 如果指定了--all，设置max_results为0
    if args.all:
        args.max_results = 0
    
    # 创建配置
    config = TenderConfig(
        keywords=args.keywords,
        output_file=args.output,
        max_results=args.max_results,
        days_back=args.days
    )
    
    if args.verbose:
        print("=" * 60)
        print("合并招标信息采集工具 (CCGP + GGZY)")
        print("=" * 60)
        print(f"\n配置信息:")
        print(f"  关键词: {', '.join(config.keywords)}")
        print(f"  时间范围: 最近 {args.days} 天")
        print(f"  输出文件: {config.output_file}")
        if config.max_results == 0:
            print(f"  最大结果数: 全部")
        else:
            print(f"  最大结果数: {config.max_results}")
        print()
    
    ccgp_results = {}
    ggzy_results = {}
    
    # 抓取CCGP
    if not args.ggzy_only:
        print("\n" + "=" * 60)
        print("【1/2】抓取中国政府采购网 (CCGP)")
        print("=" * 60)
        
        from fetchers.ccgp_fetcher import CCGPFetcher
        
        ccgp_fetcher = CCGPFetcher(config)
        
        try:
            for keyword in config.keywords:
                ccgp_results[keyword] = ccgp_fetcher.search(keyword)
                if keyword != config.keywords[-1]:
                    time.sleep(2)
        except Exception as e:
            print(f"CCGP抓取失败: {e}")
        
        ccgp_count = sum(len(tenders) for tenders in ccgp_results.values())
        print(f"\nCCGP共找到 {ccgp_count} 条记录")
    
    # 抓取GGZY
    if not args.ccgp_only:
        print("\n" + "=" * 60)
        print("【2/2】抓取全国公共资源交易平台 (GGZY)")
        print("=" * 60)
        
        try:
            from fetchers.ggzy_fetcher_playwright import GGZYFetcherPlaywright
            
            ggzy_fetcher = GGZYFetcherPlaywright(config, headless=True)
            
            try:
                ggzy_results = ggzy_fetcher.search_all_keywords()
            finally:
                ggzy_fetcher.close()
            
            ggzy_count = sum(len(tenders) for tenders in ggzy_results.values())
            print(f"\nGGZY共找到 {ggzy_count} 条记录")
            
        except ImportError as e:
            print(f"GGZY抓取失败: {e}")
            print("请确保已安装Playwright: pip install playwright && playwright install chromium")
        except Exception as e:
            print(f"GGZY抓取失败: {e}")
    
    # 导出结果
    print("\n" + "=" * 60)
    print("导出结果")
    print("=" * 60)
    
    if not ccgp_results and not ggzy_results:
        print("未找到任何招标信息")
        return
    
    output_path = export_combined_excel(ccgp_results, ggzy_results, config.output_file)
    
    print(f"\n结果已保存到: {output_path}")


if __name__ == "__main__":
    main()
